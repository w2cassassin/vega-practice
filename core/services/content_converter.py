from datetime import datetime, timedelta
from io import BytesIO
from typing import Any, Dict

from icalendar import Calendar
from openpyxl import load_workbook

from core.settings.app_config import settings

WEEKDAYS = [
    "Понедельник",
    "Вторник",
    "Среда",
    "Четверг",
    "Пятница",
    "Суббота",
    "Воскресенье",
]
COLUMN_MAPS = {
    "1": {5: "title", 6: "lesson_type", 7: "fio", 8: "room"},
    "2": {0: "title", 1: "lesson_type", 2: "fio", 3: "room"},
}


class StandardContentConverter:
    """Конвертирует файлы в стандартный JSON формат"""

    def __init__(self):
        self.supported_formats = {"xlsx": self._convert_excel, "ics": self._convert_ics}
        self.start_of_semester = None

    def convert(self, file_data: bytes, file_format: str) -> Dict[str, Any]:
        """Конвертировать файл в стандартный формат"""
        if file_format not in self.supported_formats:
            raise ValueError(f"Неподдерживаемый формат файла: {file_format}")

        return self.supported_formats[file_format](file_data)

    def _convert_excel(self, file_data: bytes) -> Dict[str, Any]:
        """Конвертировать Excel файл в стандартный формат"""
        wb = load_workbook(BytesIO(file_data))
        result = {}

        for sheet in wb.worksheets:
            block_size_first_type = 10
            block_size_second_type = 5
            max_columns = sheet.max_column

            for start_col in range(
                1, max_columns + 1, block_size_first_type + block_size_second_type
            ):
                group_name = None
                if start_col + 5 <= max_columns:
                    cell_value = sheet.cell(row=2, column=start_col + 5).value
                    if cell_value and "КМБО" in str(cell_value):
                        group_name = str(cell_value).strip()
                        column_map = COLUMN_MAPS["1"]
                        start = start_col
                        end = start_col + 9
                        self._process_block(
                            sheet, start, end, group_name, column_map, result
                        )

                if start_col + block_size_first_type <= max_columns:
                    cell_value = sheet.cell(
                        row=2, column=start_col + block_size_first_type
                    ).value
                    if cell_value and "КМБО" in str(cell_value):
                        group_name = str(cell_value).strip()
                        column_map = COLUMN_MAPS["2"]
                        start = start_col + block_size_first_type
                        end = start + 4
                        self._process_block(
                            sheet, start, end, group_name, column_map, result
                        )

        return result

    def _convert_full_name(self, full_name: str) -> str:
        """Конвертировать ФИО в сокращенный формат"""
        if not full_name:
            return ""
        parts = full_name.split()
        if len(parts) >= 3:
            return f"{parts[0]} {parts[1][0]}.{parts[2][0]}."
        return full_name

    def _process_block(self, sheet, start_col, end_col, group_name, column_map, result):
        """Обработать блок Excel и извлечь данные расписания"""
        if group_name not in result:
            result[group_name] = {day: {} for day in WEEKDAYS}

        last_teacher = None

        for row in range(4, sheet.max_row + 1):
            weekday_num = (row - 4) // 14
            if weekday_num >= len(WEEKDAYS):
                continue

            weekday = WEEKDAYS[weekday_num]
            lesson_num = str(((row - 4) % 14) // 2 + 1)
            if row % 2 == 0:
                week_type = "odd"
            else:
                week_type = "even"
            lesson_data = {
                "subject": "",
                "teacher": "",
                "room": "",
                "campus": "",
                "lesson_type": "",
                "lesson_type_id": "",
            }

            for col_idx, col in enumerate(range(start_col, end_col + 1)):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value:
                    if col_idx in column_map:
                        field = column_map[col_idx]
                        if field == "title":
                            lesson_data["subject"] = str(cell_value)
                        elif field == "fio":
                            lesson_data["teacher"] = str(cell_value)
                            # last_teacher = lesson_data["teacher"]
                        elif field == "lesson_type":
                            for type_prefix, type_id in settings.LESSON_TYPES.items():
                                if str(cell_value).strip() == type_prefix:
                                    lesson_data["lesson_type"] = type_prefix
                                    lesson_data["lesson_type_id"] = type_id
                                    break
                        elif field == "room":
                            room_parts = (
                                str(cell_value)
                                .replace("ауд. ", "")
                                .replace("комп. ", "")
                                .split()
                            )
                            lesson_data["room"] = room_parts[0] if room_parts else ""
                            lesson_data["campus"] = (
                                room_parts[1].replace("(", "").replace(")", "")
                                if len(room_parts) > 1
                                else ""
                            )
            if len(lesson_data["subject"]) < 3:
                continue
            if not lesson_data["teacher"] and lesson_data["subject"] and last_teacher:
                lesson_data["teacher"] = last_teacher

            if any(value for key, value in lesson_data.items() if key != "campus"):
                result[group_name][weekday][lesson_num] = result[group_name][
                    weekday
                ].get(lesson_num, {})
                result[group_name][weekday][lesson_num][week_type] = lesson_data

    def _convert_ics(self, file_data: bytes) -> Dict[str, Any]:
        """Конвертировать ICS файл в стандартный формат"""
        cal = Calendar.from_ical(file_data)
        result = {}
        group_name = str(cal.get("X-WR-CALNAME"))
        temp_events = []
        event_dates = []

        # Сначала собираем все даты
        for component in cal.walk():
            if component.name == "VEVENT":
                start_date = component.get("dtstart").dt
                if isinstance(start_date, datetime):
                    event_dates.append(start_date)

        if not event_dates:
            # Если нет событий, возвращаем пустой результат
            return {}

        # Определим дату начала семестра как понедельник той недели, когда было первое событие
        earliest_date = min(event_dates)
        # Найдем понедельник той же недели
        start_of_week = earliest_date - timedelta(
            days=earliest_date.weekday()
        )  # weekday() дает 0 для понедельника
        self.start_of_semester = start_of_week

        # Теперь повторно пройдемся по событиям
        for component in cal.walk():
            if component.name == "VEVENT":
                summary = str(component.get("summary", ""))
                description = str(component.get("description", ""))

                start_date = component.get("dtstart").dt
                weekday = WEEKDAYS[start_date.weekday()]
                lesson_num = str(self._get_lesson_number(start_date))

                if lesson_num == "0":
                    continue

                subject = summary
                lesson_type = None
                lesson_type_id = None
                for type_prefix, type_id in settings.LESSON_TYPES.items():
                    if summary.startswith(type_prefix):
                        subject = summary[len(type_prefix) :].strip()
                        lesson_type = type_prefix
                        lesson_type_id = type_id
                        break

                location = str(component.get("location", ""))
                room, campus = self._extract_room_campus(location)
                teacher = self._extract_teacher(description)

                # Проверим правило повторения
                rrule = component.get("RRULE")
                parity = None
                if rrule:
                    interval = rrule.get("INTERVAL", [1])[0]
                    if interval == 2:
                        # Если мероприятие повторяется каждые две недели
                        week_number = self._get_week_number(start_date)
                        parity = self._get_week_parity(week_number)

                data = {
                    "subject": subject,
                    "teacher": teacher,
                    "room": room,
                    "campus": campus,
                    "lesson_type": lesson_type,
                    "lesson_type_id": lesson_type_id,
                }

                temp_events.append(
                    {
                        "group": group_name,
                        "weekday": weekday,
                        "lesson_num": lesson_num,
                        "parity": parity,
                        "data": data,
                    }
                )

        # Сформируем итоговую структуру
        for event in temp_events:
            group = event["group"]
            weekday = event["weekday"]
            lesson_num = event["lesson_num"]
            parity = event["parity"]  # будет None, если повтор каждую неделю

            if group not in result:
                result[group] = {day: {} for day in WEEKDAYS}

            if parity:
                if lesson_num not in result[group][weekday]:
                    result[group][weekday][lesson_num] = {}
                result[group][weekday][lesson_num][parity] = event["data"]
            else:
                # Нет четности
                result[group][weekday][lesson_num] = event["data"]

        return result

    def _extract_room_campus(self, location: str) -> tuple:
        """Извлечь аудиторию и кампус из строки"""
        parts = location.split(" ")
        room = parts[0].strip() if parts else ""
        campus = (
            parts[1].strip().replace("(", "").replace(")", "") if len(parts) > 1 else ""
        )
        return room, campus

    def _extract_teacher(self, description: str) -> str:
        """Извлечь имя преподавателя из описания и конвертировать в сокращенный формат"""
        if "Преподаватель:" in description:
            full_name = description.split("Преподаватель:")[1].split("\n")[0].strip()
            return self._convert_full_name(full_name)
        return ""

    def _get_lesson_number(self, start_time) -> int:
        """Конвертировать время начала в номер пары"""
        time_to_lesson = {
            "09:00": 1,
            "10:40": 2,
            "12:40": 3,
            "14:20": 4,
            "16:20": 5,
            "18:00": 6,
            "19:40": 7,
        }
        time_str = start_time.strftime("%H:%M")
        return time_to_lesson.get(time_str, 0)

    def _get_week_number(self, date: datetime) -> int:
        """Определить номер недели относительно start_of_semester"""
        delta = date.date() - self.start_of_semester.date()
        # Номер недели (начиная с 1)
        week_number = delta.days // 7 + 1
        return week_number

    def _get_week_parity(self, week_number: int) -> str:
        """Вернуть 'even' или 'odd' в зависимости от номера недели"""
        return "even" if week_number % 2 == 0 else "odd"
