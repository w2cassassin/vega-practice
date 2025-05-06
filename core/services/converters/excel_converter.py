from io import BytesIO
from typing import Dict
from openpyxl import load_workbook

from core.settings.app_config import settings
from core.schemas.schedule import LessonData, ScheduleResult
from core.utils.maps import WEEKDAYS, COLUMN_MAPS
from core.services.converters.base_converter import BaseConverter


class ExcelConverter(BaseConverter):
    """Конвертер для Excel-файлов расписания"""

    def convert(self, file_data: bytes) -> ScheduleResult:
        """Конвертировать Excel-файл в стандартный формат"""
        wb = load_workbook(BytesIO(file_data))

        results: Dict[str, ScheduleResult] = {}

        for sheet in wb.worksheets:
            self._process_sheet(sheet, results)

        if not results:
            return ScheduleResult()

        return next(iter(results.values()))

    def _process_sheet(self, sheet, results: Dict[str, ScheduleResult]):
        """Обработать лист Excel"""
        block_size_first_type = 10
        block_size_second_type = 5
        max_columns = sheet.max_column

        for start_col in range(
            1, max_columns + 1, block_size_first_type + block_size_second_type
        ):
            if start_col + 5 <= max_columns:
                cell_value = sheet.cell(row=2, column=start_col + 5).value
                if cell_value and "КМБО" in str(cell_value):
                    group_name = str(cell_value).strip()
                    if group_name not in results:
                        results[group_name] = ScheduleResult(group_name=group_name)

                    self._process_block(
                        sheet,
                        start_col,
                        start_col + 9,
                        results[group_name],
                        COLUMN_MAPS["1"],
                    )

            if start_col + block_size_first_type <= max_columns:
                cell_value = sheet.cell(
                    row=2, column=start_col + block_size_first_type
                ).value
                if cell_value and "КМБО" in str(cell_value):
                    group_name = str(cell_value).strip()
                    if group_name not in results:
                        results[group_name] = ScheduleResult(group_name=group_name)

                    self._process_block(
                        sheet,
                        start_col + block_size_first_type,
                        start_col + block_size_first_type + 4,
                        results[group_name],
                        COLUMN_MAPS["2"],
                    )

    def _process_block(
        self, sheet, start_col, end_col, result: ScheduleResult, column_map
    ):
        """Обработать блок Excel и извлечь данные расписания"""
        last_teacher = None

        for row in range(4, sheet.max_row + 1):
            weekday_num = (row - 4) // 14
            if weekday_num >= len(WEEKDAYS):
                continue

            weekday = WEEKDAYS[weekday_num]
            lesson_num = str(((row - 4) % 14) // 2 + 1)
            is_even_row = row % 2 == 0

            week_numbers = []
            if is_even_row:  # нечетные недели (odd)
                week_numbers = list(
                    range(1, self.total_weeks + 1, 2)
                )  # 1, 3, 5, 7, ...
            else:  # четные недели (even)
                week_numbers = list(
                    range(2, self.total_weeks + 1, 2)
                )  # 2, 4, 6, 8, ...

            lesson_data = LessonData()

            for col_idx, col in enumerate(range(start_col, end_col + 1)):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value:
                    if col_idx in column_map:
                        field = column_map[col_idx]
                        if field == "title":
                            lesson_data.subject = str(cell_value)
                        elif field == "fio":
                            lesson_data.teacher = str(cell_value)
                        elif field == "lesson_type":
                            for type_prefix, type_id in settings.LESSON_TYPES.items():
                                if str(cell_value).strip() == type_prefix:
                                    lesson_data.lesson_type = type_prefix
                                    lesson_data.lesson_type_id = type_id
                                    break
                        elif field == "room":
                            room_parts = (
                                str(cell_value)
                                .replace("ауд. ", "")
                                .replace("комп. ", "")
                                .split()
                            )
                            lesson_data.room = room_parts[0] if room_parts else ""
                            lesson_data.campus = (
                                room_parts[1].replace("(", "").replace(")", "")
                                if len(room_parts) > 1
                                else ""
                            )

            if len(lesson_data.subject) < 3:
                continue

            if not lesson_data.teacher and lesson_data.subject and last_teacher:
                lesson_data.teacher = last_teacher
            elif lesson_data.teacher:
                last_teacher = lesson_data.teacher

            if any(
                getattr(lesson_data, attr) for attr in ["subject", "teacher", "room"]
            ):
                for week_number in week_numbers:
                    week_lesson_data = LessonData(
                        subject=lesson_data.subject,
                        teacher=lesson_data.teacher,
                        room=lesson_data.room,
                        campus=lesson_data.campus,
                        lesson_type=lesson_data.lesson_type,
                        lesson_type_id=lesson_data.lesson_type_id,
                    )

                    self._add_lesson_to_schedule(
                        result, week_number, weekday, lesson_num, week_lesson_data
                    )
