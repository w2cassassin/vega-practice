from copy import deepcopy
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

from core.services.base import BaseDocumentService

COLUMN_MAPS = {
    "1": {5: "title", 7: "fio", 8: "room"},
    "2": {0: "title", 2: "fio", 3: "room"},
}
WEEKDAYS = [
    "Понедельник",
    "Вторник",
    "Среда",
    "Четверг",
    "Пятница",
    "Суббота",
    "Воскресенье",
]


class ExcelService(BaseDocumentService):
    def __init__(self):
        self.workbook = None

    def load(self, file) -> None:
        self.workbook = load_workbook(file)

    def update(self, params: dict) -> None:
        if not self.workbook:
            raise ValueError("Excel файл не загружен.")
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.hyperlink and cell.hyperlink.display in params:
                        value = params.get(cell.hyperlink.display)

                        if cell.hyperlink.location:
                            location_sheet, location_cell = (
                                cell.hyperlink.location.split("!")
                            )
                            location_sheet = location_sheet.replace("'", "")
                            location_sheet = self.workbook[location_sheet]
                            start_cell = location_sheet[location_cell]

                            if isinstance(
                                value, list
                            ):  # Если значение словаря - список из списков
                                for i, sublist in enumerate(
                                    value
                                ):  # Проход по спискам списка
                                    for j, item in enumerate(
                                        sublist
                                    ):  # Проход по значениям из вложенного списка
                                        target_cell_col = start_cell.column + j
                                        target_cell_row = start_cell.row + i
                                        col_letter = get_column_letter(target_cell_col)
                                        self._write_value_to_cell(
                                            location_sheet,
                                            f"{col_letter}{target_cell_row}",
                                            item,
                                        )

                            else:
                                self._write_value_to_cell(sheet, location_cell, value)

    def to_json(self, sheet_name: str = None, range: str = None) -> dict:
        if not self.workbook:
            raise ValueError("Excel файл не загружен.")
        sheets = self.workbook.worksheets
        if sheet_name and sheet_name in self.workbook.sheetnames:
            sheets = [self.workbook[sheet_name]]

        data = {}

        for sheet in sheets:
            base_range = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
            if range:
                base_range = range
            sheet_data = {"cells": {}, "merged": []}

            merged_ranges = [str(merge) for merge in sheet.merged_cells.ranges]
            sheet_data["merged"] = merged_ranges

            cells_range = sheet[base_range]
            for row in cells_range:
                for cell in row:
                    if cell.value is not None:
                        cell_letter = get_column_letter(cell.column)
                        cell_address = f"{cell_letter}{cell.row}"

                        cell_format = self._extract_cell_format(cell)

                        sheet_data["cells"][cell_address] = {
                            "value": cell.value,
                            "format": cell_format,
                        }

            if sheet_data["cells"]:
                data[sheet.title] = sheet_data

        return data

    def _extract_cell_format(self, cell):
        """Извлекает информацию о форматировании из ячейки."""
        format_dict = {}

        font = cell.font
        if font:
            if font.color:
                color = font.color
                if color.type == "rgb" and color.rgb:
                    format_dict["textcolor"] = color.rgb[2:]
            if font.name:
                format_dict["fontname"] = font.name
            if font.size:
                format_dict["fontsize"] = font.size
            if font.bold:
                format_dict["bold"] = font.bold
            if font.italic:
                format_dict["italic"] = font.italic
            if font.underline and font.underline != "none":
                format_dict["underline"] = True
            if font.strike:
                format_dict["strikethrough"] = font.strike

        fill = cell.fill
        if fill and fill.patternType == "solid":
            fgColor = fill.fgColor
            if fgColor:
                if fgColor.type == "rgb" and fgColor.rgb:
                    format_dict["fillcolor"] = fgColor.rgb[-6:]

        alignment = cell.alignment
        if alignment:
            if alignment.horizontal:
                format_dict["align"] = alignment.horizontal
            if alignment.vertical:
                format_dict["valign"] = alignment.vertical

        return format_dict

    def from_json(self, data: dict) -> None:
        if not self.workbook:
            raise ValueError("Excel файл не загружен.")

        for sheet_name, sheet_data in data.items():
            if sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                sheet.delete_rows(1, sheet.max_row)
            else:
                sheet = self.workbook.create_sheet(sheet_name)

            merged_ranges = sheet_data.get("merged", [])
            for merge_range in merged_ranges:
                self._merge_cells_with_data(sheet, merge_range)

            cells_data = sheet_data.get("cells", {})
            for cell_address, cell_info in cells_data.items():
                value = cell_info.get("value")
                format_dict = cell_info.get("format", {})

                self._write_value_to_cell(sheet, cell_address, value)

                cell = sheet[cell_address]
                self._apply_formatting(cell, format_dict)

    def update_with_blocks(self, data: dict) -> None:
        if not self.workbook:
            raise ValueError("Excel файл не загружен.")

        original_workbook = deepcopy(self.workbook)

        blocks = data.get("blocks", [])
        newpage = data.get("newpage", True)
        merged_ranges = data.get("merged", [])
        format_data = data.get("format", {})

        sheet = self.workbook.active
        original_sheet = original_workbook.active

        if newpage:
            sheet_counter = 1
            for block in blocks:
                if sheet_counter > 1:
                    new_sheet = self.workbook.copy_worksheet(sheet)
                    new_sheet.title = f"{sheet.title}_{sheet_counter}"
                else:
                    new_sheet = sheet

                new_sheet.delete_rows(1, new_sheet.max_row)
                self._copy_sheet_content(original_sheet, new_sheet)

                self._process_block(new_sheet, block, merged_ranges, format_data)

                sheet_counter += 1
        else:
            for i, block in enumerate(blocks):
                self._copy_sheet_content(original_sheet, sheet)
                num_rows = original_sheet.max_row

                self._process_block(sheet, block, merged_ranges, format_data)

                if i < len(blocks) - 1:
                    sheet.insert_rows(1, num_rows + 1)

    def _process_block(self, sheet, block, merged_ranges, format_data):

        for cell_address, value in block.items():
            self._write_value_to_cell(sheet, cell_address, value)

        for merge_range in merged_ranges:
            self._merge_cells_with_data(sheet, merge_range)

        for cell_range, format_dict in format_data.items():
            if ":" in cell_range:
                for row in sheet[cell_range]:
                    for cell in row:
                        self._apply_formatting(cell, format_dict)
            else:
                self._apply_formatting(sheet[cell_range], format_dict)

    def _apply_formatting(self, cell, format_dict):
        font_args = {}

        if "textcolor" in format_dict:
            font_args["color"] = format_dict["textcolor"]

        if "fontname" in format_dict:
            font_args["name"] = format_dict["fontname"]

        if "fontsize" in format_dict:
            font_args["size"] = format_dict["fontsize"]

        if "bold" in format_dict:
            font_args["bold"] = format_dict["bold"]

        if "italic" in format_dict:
            font_args["italic"] = format_dict["italic"]

        if "underline" in format_dict:
            underline_value = format_dict["underline"]
            font_args["underline"] = "single" if underline_value else None

        if "strikethrough" in format_dict:
            font_args["strike"] = format_dict["strikethrough"]

        if font_args:
            cell.font = Font(**font_args)

        if "fillcolor" in format_dict:
            cell.fill = PatternFill(
                start_color=format_dict["fillcolor"],
                end_color=format_dict["fillcolor"],
                fill_type="solid",
            )

        alignment_args = {}

        if "align" in format_dict:
            alignment_args["horizontal"] = format_dict["align"]

        if "valign" in format_dict:
            alignment_args["vertical"] = format_dict["valign"]

        if alignment_args:
            cell.alignment = Alignment(**alignment_args)

    def _copy_sheet_content(self, source_sheet, target_sheet):
        for row in source_sheet.iter_rows():
            for cell in row:
                target_cell = target_sheet.cell(row=cell.row, column=cell.column)
                if isinstance(target_cell, MergedCell):
                    continue

                target_cell.value = cell.value

                if cell.has_style:
                    target_cell._style = cell._style
                if cell.hyperlink:
                    target_cell._hyperlink = cell.hyperlink
                if cell.comment:
                    target_cell.comment = cell.comment

        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))

    def _write_value_to_cell(self, sheet, cell_address, value):
        target_cell = sheet[cell_address]
        if "\n" in str(value):
            target_cell.alignment = target_cell.alignment.copy(wrap_text=True)
        merged_range_to_recreate = None
        if isinstance(target_cell, MergedCell):
            for merged_range in sheet.merged_cells.ranges:
                if cell_address in merged_range:
                    merged_range_to_recreate = str(merged_range)
                    sheet.unmerge_cells(merged_range_to_recreate)
                    break
        sheet[cell_address] = value
        if merged_range_to_recreate:
            sheet.merge_cells(merged_range_to_recreate)

    def _merge_cells_with_data(self, sheet, merge_range):
        min_col, min_row, max_col, max_row = range_boundaries(merge_range)

        merged_data = []
        for row in sheet.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        ):
            for cell in row:
                if cell.value:
                    merged_data.append(str(cell.value))

        combined_value = "\n".join(merged_data)
        cell_address = f"{get_column_letter(min_col)}{min_row}"
        self._write_value_to_cell(sheet, cell_address, combined_value)
        sheet.merge_cells(merge_range)

    def save_to_bytes(self) -> BytesIO:
        if not self.workbook:
            raise ValueError("Excel файл не загружен.")
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output

    def save_to_file(self, file_path: str) -> None:
        if self.workbook:
            self.workbook.save(file_path)
        else:
            raise ValueError("Excel файл не загружен.")


class ExcelCompareService(ExcelService):

    def compare(self, file_to_compare: BytesIO) -> None:
        """Сравнивает два Excel файла и перемещает блоки с КМБО в начало."""
        if not self.workbook:
            raise ValueError("Основной Excel файл не загружен.")

        # Загружаем второй файл для сравнения
        compare_workbook = load_workbook(file_to_compare)

        # Проход по всем листам
        for sheet_name in self.workbook.sheetnames:
            sheet1 = self.workbook[sheet_name]
            sheet2 = compare_workbook[sheet_name]

            # Выделяем блоки с группами КМБО и сравниваем
            kmbo_blocks = self._get_kmbo_blocks(sheet1)
            kmbo_blocks_compare = self._get_kmbo_blocks(sheet2)

            # Сравниваем блоки и подсвечиваем различия
            compare_data = self._compare_and_highlight(
                sheet1, kmbo_blocks, sheet2, kmbo_blocks_compare
            )
            self._move_blocks_to_start(sheet1, kmbo_blocks)
            return compare_data

    def _get_kmbo_blocks(self, sheet):
        """Извлекает блоки с КМБО для сравнения."""
        kmbo_blocks = []
        max_columns = sheet.max_column

        block_size_first_type = 10  # Для блока первого типа (A-J)
        block_size_second_type = 5  # Для блока второго типа (K-O)

        for start_col in range(
            1, max_columns + 1, block_size_first_type + block_size_second_type
        ):
            # Обрабатываем блоки первого типа (A-J)
            first_block_range = self._identify_block_range(sheet, start_col, 1)
            if first_block_range:
                kmbo_blocks.append(first_block_range)

            # Обрабатываем блоки второго типа (K-O)
            second_block_range = self._identify_block_range(
                sheet, start_col + block_size_first_type, 2
            )
            if second_block_range:
                kmbo_blocks.append(second_block_range)

        return kmbo_blocks

    def _identify_block_range(self, sheet, start_col, block_type):
        """Определяет диапазон блока в зависимости от его типа (первый или второй)."""
        if block_type == 1:
            # Название группы находится в 6-м столбце блока
            group_col = start_col + 5
            if sheet.cell(row=2, column=group_col).value and "КМБО" in str(
                sheet.cell(row=2, column=group_col).value
            ):
                return (start_col, start_col + 9)
        elif block_type == 2:
            # Название группы находится в 1-м столбце блока
            group_col = start_col
            if sheet.cell(row=2, column=group_col).value and "КМБО" in str(
                sheet.cell(row=2, column=group_col).value
            ):
                return (start_col, start_col + 4)
        return None

    def _compare_and_highlight(self, sheet1, blocks1, sheet2, blocks2):
        """Сравнивает блоки и подсвечивает различия."""
        red_fill = PatternFill(
            start_color="FF0000", end_color="FF0000", fill_type="solid"
        )
        compare_data = {"title": 0, "fio": 0, "room": 0, "campus": 0, "changes": {}}
        for block1, block2 in zip(blocks1, blocks2):
            index = 0
            if block1[1] - block1[0] == 9:
                column_map = COLUMN_MAPS["1"]
                group_name = sheet1.cell(row=2, column=block1[0] + 5).value
            else:
                column_map = COLUMN_MAPS["2"]
                group_name = sheet1.cell(row=2, column=block1[0]).value
            print(group_name)
            for col in range(block1[0], block1[1] + 1):
                for row in range(1, sheet1.max_row + 1):
                    weekday_num = (row - 4) // 14
                    weekday = (
                        WEEKDAYS[weekday_num] if weekday_num < len(WEEKDAYS) else "-"
                    )
                    lesson_num = ((row - 4) % 14) // 2 + 1
                    cell1 = sheet1.cell(row=row, column=col)
                    cell2 = sheet2.cell(row=row, column=col)
                    if cell1.value != cell2.value:
                        if group_name not in compare_data["changes"]:
                            compare_data["changes"][group_name] = {}
                        if weekday not in compare_data["changes"][group_name]:
                            compare_data["changes"][group_name][weekday] = {}
                        if (
                            lesson_num
                            not in compare_data["changes"][group_name][weekday]
                        ):
                            compare_data["changes"][group_name][weekday][
                                lesson_num
                            ] = []
                        compare_data["changes"][group_name][weekday][lesson_num].append(
                            f"{cell2.value} -> {cell1.value}"
                        )
                        if index in column_map:
                            column_name = column_map[index]
                            if column_name == "room" and (
                                cell1.value in cell2.value or cell2.value in cell1.value
                            ):
                                compare_data["campus"] += 1
                            else:
                                compare_data[column_name] += 1
                        cell1.fill = red_fill
                index += 1
        return compare_data

    def _move_blocks_to_start(self, sheet, blocks):
        """Перемещает блоки с КМБО в начало таблицы и удаляет все, что справа."""
        inserted_cols = 1
        for block in blocks:
            start_col, end_col = block[0], block[1]
            sheet.move_range(
                f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}{sheet.max_row}",
                cols=-start_col + inserted_cols,
            )
            inserted_cols += end_col - start_col + 1

        last_col = inserted_cols - 1
        sheet.delete_cols(last_col + 1, sheet.max_column - last_col)
